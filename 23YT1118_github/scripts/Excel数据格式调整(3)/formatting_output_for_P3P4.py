import pandas as pd
import os
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelFormat:
    def __init__(self, file_path, sheet_name, start_col) -> None:
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.start_col = start_col       
        
    
    def main(self):
        # 读取Excel文件
        df = pd.read_excel(self.file_path, header=None, sheet_name=self.sheet_name)
        # 定义条件函数
        condition = lambda row : 'All' in row.values
        # 标记满足条件的行
        mask = df.apply(condition, axis=1)
        # 创建边框样式
        border_style_common = Border(top=Side(border_style="thin", color="000000")
                                    , left=Side(border_style="thin", color="000000")
                                    , bottom=Side(border_style="thin", color="000000")
                                    , right=Side(border_style="thin", color="000000"))
        border_style_row = Border(top=Side(border_style="thin", color="000000")
                                    , left=Side(border_style="thin", color="000000")
                                    , bottom=Side(border_style="thick", color="000000")
                                    , right=Side(border_style="thin", color="000000"))
        border_style_col = Border(top=Side(border_style="thin", color="000000")
                                    , left=Side(border_style="thin", color="000000")
                                    , bottom=Side(border_style="thin", color="000000")
                                    ,right=Side(border_style="thick", color="000000"))
        border_style_row_col = Border(top=Side(border_style="thin", color="000000")
                                    , left=Side(border_style="thin", color="000000")
                                    , bottom=Side(border_style="thick", color="000000")
                                    , right=Side(border_style="thick", color="000000"))
        # 保存修改后的数据回原文件
        book = load_workbook(self.file_path)
        writer = pd.ExcelWriter(self.file_path, engine='openpyxl')
        writer.book = book
        # 获取工作表对象
        ws = writer.book[self.sheet_name]
        # 首行和首列窗口冻结
        ws.freeze_panes = 'B2'
        # 开启筛选
        ws.auto_filter.ref = ws.dimensions
        # 将满足条件的行的下边框设置为深色
        for row_num, is_match in enumerate(mask, start=1):
            for col_num in range(1, df.shape[1]+1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border_style_common
        for row_num, is_match in enumerate(mask, start=1):
            if is_match:
                for col_num in range(1, df.shape[1]+1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border_style_row
        for row_num, is_match in enumerate(mask, start=1):
            for col_num in range(1, df.shape[1]+1):
                if (col_num > self.start_col and (col_num - self.start_col) % 4 == 0) or col_num == self.start_col:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = border_style_col
                    if is_match:
                        cell.border = border_style_row_col
        for col_num in range(1, df.shape[1]+1):
            cell = ws.cell(row=1, column=col_num)
            # cell.fill = fill_style_1
            cell.alignment = Alignment(vertical='top', horizontal='center', text_rotation=90)
            if (col_num > self.start_col and (col_num - self.start_col) % 4 == 0) or col_num == self.start_col :
                cell.border = border_style_row_col
            else:
                cell.border = border_style_row
        # 表格填充颜色
        fill_style_1 = PatternFill(start_color="a3c99e", end_color="FF0000", fill_type="solid")  # 红色填充
        fill_style_2 = PatternFill(start_color="daeef3", end_color="00FFFF", fill_type="solid")  # 蓝色填充
        fill_style_3 = PatternFill(start_color="eeece1", end_color="00FF80", fill_type="solid")  # 绿色填充
        color_list = [fill_style_2, fill_style_3]
        for row_num, is_match in enumerate(mask, start=1):
            if row_num == len(mask) + 1:
                continue
            select_color_num = 0
            select_color = color_list[select_color_num]
            color_time = 0
            for col_num in range(1, df.shape[1]+1):
                if col_num > self.start_col:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.fill = select_color
                color_time += 1
                if (col_num > self.start_col and ((col_num - self.start_col) % 4 == 0) and color_time != 0):
                    select_color_num += 1
                    select_color = color_list[select_color_num % 2]
        # 设置列宽自适应
        pixel_width_to_character_width = lambda pixel_width: (pixel_width - 5) / 7  # 像素数量转化为字符长度
        for index, column in enumerate(ws.columns):
            column_letter = get_column_letter(column[0].column)
            if self.start_col <= index <= (self.start_col + 3):
                adjusted_width = pixel_width_to_character_width(61*0.60)
            else:
                adjusted_width = pixel_width_to_character_width(61*0.55)
            ws.column_dimensions[column_letter].width = adjusted_width
        # 为特定列执行列宽
        ws.column_dimensions["A"].width = pixel_width_to_character_width(69)
        ws.column_dimensions["B"].width = pixel_width_to_character_width(33)
        ws.column_dimensions["C"].width = pixel_width_to_character_width(33)
        ws.column_dimensions["D"].width = pixel_width_to_character_width(50)
        ws.column_dimensions["E"].width = pixel_width_to_character_width(63)
        ws.column_dimensions["F"].width = pixel_width_to_character_width(40)
        # 设置行高
        ws.row_dimensions[1].height = 80
        df.to_excel(writer, index=False, header=False, sheet_name=self.sheet_name)
        writer.save()
        writer.close()


class Excel:
    def __init__(self, file_path="./data/test.xlsx", result_folder="./result/", sheet_name="所有熟期", start_col=4) -> None:
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.result_file = result_folder + os.path.basename(self.file_path).split(".")[0] + "_result.xlsx"
        self.start_col = start_col
        # P1和P3P4用
        self.sheet_selection_colum = "AOA"  # 划分sheet表单的时候用的列名
        self.sheet_name_to_item_dict = {
                                        "EMSP": ["EMSP"]
                                        , "SP": ["LMSP", "MMSP", "SP"]
                                        , "SU": ["NCSU", "MCSU", "SCSU", "SU"]
                                        , "SWCN": ["SWCN"]
                                        }  # 每个sheet中需要包含的样本信息（sheet_selection_colum中的某些样本名称）
        # TC1/TC2用
        # self.sheet_selection_colum = "Station"  # 划分sheet表单的时候用的列名
        # self.sheet_name_to_item_dict = {
        #                                 "HRB": ["HRB"]
        #                                 , "SP": ["CC", "NW", "SP"]
        #                                 , "SU": ["SU", "JN", "XX"]
        #                                 , "SW": ["SW"]
        #                                 }  # 每个sheet中需要包含的样本信息（sheet_selection_colum中的某些样本名称）
        if not os.path.exists(result_folder):
            os.makedirs(result_folder)
    
    def get_file_content(self):
        print("Reading file...")
        file_content_df = pd.read_excel(self.file_path, header=0, sheet_name=self.sheet_name)
        return file_content_df
    
    def split_sample(self):
        """1103result_by_23TC1_v1(1)-Sort-Y1.xlsx文件用"""
        file_content_df = self.get_file_content()
        writer = pd.ExcelWriter(self.result_file)
        for sheet_name in tqdm(self.sheet_name_to_item_dict, desc="Sheet split process"):
            sheet_value_list = self.sheet_name_to_item_dict[sheet_name]
            sheet = file_content_df[file_content_df[self.sheet_selection_colum].isin(sheet_value_list)]
            sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()
        writer.close()
        return None
    
    def change_excel_format(self):
        for sheet_name in tqdm(self.sheet_name_to_item_dict.keys(), desc="Sheet format process"):
            excelFormat = ExcelFormat(file_path=self.result_file, sheet_name=sheet_name, start_col=self.start_col)
            excelFormat.sheet_name = sheet_name
            excelFormat.main()
    
    def main(self):
        self.split_sample()
        self.change_excel_format()
        
        
if __name__ == "__main__":
    os.chdir("E:/23YT1103/Excel数据格式调整(3)")
    # Excel格式调整
    # excelFormat = ExcelFormat(file_path="./test_result.xlsx", sheet_name="SP")
    # excelFormat.main()
    # Excel的sheet划分及格式调整（调用ExcelFormat对象）
    excel = Excel(file_path='./data/2023年-P3P4试验-杂交种产量及病害分析-20231104.xlsx', result_folder='./result/', sheet_name="Sheet 1", start_col=6)
    excel.main()